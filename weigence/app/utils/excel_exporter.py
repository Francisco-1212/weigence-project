"""
Sistema Weigence - Exportador de Excel Profesional
Genera archivos Excel estéticos con formato empresarial
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.utils.vencimiento_helper import VencimientoHelper


class ExcelExporter:
    """Clase para generar reportes Excel profesionales"""
    
    # Colores corporativos Weigence
    COLOR_HEADER = "1E3A8A"      # Azul oscuro
    COLOR_SUBTITLE = "3B82F6"    # Azul medio
    COLOR_SUCCESS = "10B981"     # Verde
    COLOR_WARNING = "F59E0B"     # Amarillo
    COLOR_DANGER = "DC2626"      # Rojo
    COLOR_INFO = "3B82F6"        # Azul
    COLOR_GRAY = "6B7280"        # Gris
    COLOR_ALTERNADO = "F3F4F6"   # Gris claro para filas alternas
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
    def _aplicar_estilo_header(self, cell, color="1E3A8A"):
        """Aplica estilo de encabezado principal"""
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self._crear_borde()
        
    def _aplicar_estilo_celda(self, cell, bold=False, color_fondo=None, color_texto=None):
        """Aplica estilo a una celda de datos"""
        cell.font = Font(name='Calibri', size=10, bold=bold, color=color_texto or "000000")
        if color_fondo:
            cell.fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = self._crear_borde(style='thin')
        
    def _crear_borde(self, style='medium'):
        """Crea un borde para las celdas"""
        side = Side(style=style, color="000000")
        return Border(left=side, right=side, top=side, bottom=side)
    
    def _ajustar_anchos_columna(self, min_width=8, max_width=50):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def _agregar_encabezado_principal(self, titulo, total_registros, filtros=None):
        """Agrega encabezado principal al reporte"""
        col_count = self.ws.max_column or 9
        row = 1
        
        # Fila 1: Logo/Marca en columna K
        cell_marca = self.ws['K1']
        cell_marca.value = f"🏢 SISTEMA WEIGENCE - 📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        cell_marca.font = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
        cell_marca.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell_marca.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.ws.row_dimensions[row].height = 30
        self.ws.column_dimensions['K'].width = 25
        
        # Fila 2: Título principal en columna K
        row += 1
        cell_title = self.ws['K2']
        cell_title.value = titulo
        cell_title.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        cell_title.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        cell_title.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.ws.row_dimensions[row].height = 40
        
        # Fila 3: Información del total de registros en columna K
        row += 1
        total_cell = self.ws['K3']
        total_cell.value = f"📊 Total de registros: {total_registros}"
        total_cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
        total_cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        total_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.ws.row_dimensions[row].height = 30
        
        # Filtros aplicados en columna K
        if filtros:
            filtros_text = " | ".join([f"{k.capitalize()}: {v}" for k, v in filtros.items() if v])
            if filtros_text:
                row += 1
                filtro_cell = self.ws[f'K{row}']
                filtro_cell.value = f"🔍 Filtros: {filtros_text}"
                filtro_cell.font = Font(name='Calibri', size=9, italic=True, color="FFFFFF")
                filtro_cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                filtro_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                self.ws.row_dimensions[row].height = 25
        
        return 1  # Retorna la fila donde empezar el contenido (fila 1 para que los headers empiecen ahí)
    
    def _agregar_pie_pagina(self):
        """Agrega pie de página al reporte"""
        row = self.ws.max_row + 2
        col_count = self.ws.max_column
        
        # Línea decorativa antes del footer
        self.ws.merge_cells(f'A{row}:{get_column_letter(col_count)}{row}')
        linea_cell = self.ws[f'A{row}']
        linea_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        self.ws.row_dimensions[row].height = 3
        
        row += 1
        self.ws.merge_cells(f'A{row}:{get_column_letter(col_count)}{row}')
        footer_cell = self.ws[f'A{row}']
        footer_cell.value = "💼 Sistema Weigence - Gestión Inteligente de Inventario | 📄 Generado automáticamente"
        footer_cell.font = Font(name='Calibri', size=9, italic=True, color="6B7280", bold=True)
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')
        footer_cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        self.ws.row_dimensions[row].height = 25
    
    def exportar_inventario(self, productos, filtros=None):
        """Exporta inventario a Excel con formato profesional"""
        self.ws.title = "Inventario"
        
        # Encabezado
        row = self._agregar_encabezado_principal(
            "SISTEMA WEIGENCE - REPORTE DE INVENTARIO",
            len(productos),
            filtros
        )
        
        # Encabezados de columnas
        headers = ['ID', 'Nombre', 'Categoría', 'Stock', 'Peso (kg)', 'Estante', 'Estado', 'F. Elaboración', 'F. Vencimiento', 'Días Rest.', 'Fecha Ingreso']
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_num)
            cell.value = header
            self._aplicar_estilo_header(cell, self.COLOR_SUBTITLE)
        
        row += 1
        
        # Datos de productos
        for idx, producto in enumerate(productos):
            color_fondo = self.COLOR_ALTERNADO if idx % 2 == 0 else None
            
            # Determinar color según stock
            stock = producto.get('stock', 0)
            
            # Determinar estado basado en stock (misma lógica que inventario.js: stock < 10)
            if stock == 0:
                estado = "Sin Stock"
                color_fondo = "FEE2E2"  # Rojo claro
            elif 0 < stock < 10:
                estado = "Bajo Stock"
                color_fondo = "FEF3C7"  # Amarillo claro
            else:
                estado = "Normal"
            
            # Verificar vencimiento (prioridad sobre stock)
            fecha_venc = producto.get('fecha_vencimiento')
            estado_vencimiento = None
            dias_restantes = None
            
            if fecha_venc:
                estado_venc = VencimientoHelper.obtener_estado_vencimiento(fecha_venc)
                dias_restantes = estado_venc['dias_restantes']
                estado_vencimiento = estado_venc['estado']
                
                # Vencimiento tiene prioridad sobre stock en el color
                if estado_vencimiento == 'vencido' or estado_vencimiento == 'vence_hoy':
                    color_fondo = "FEE2E2"  # Rojo más intenso
                elif estado_vencimiento == 'critico':
                    color_fondo = "FED7AA"  # Naranja
                elif estado_vencimiento == 'proximo':
                    color_fondo = "FEF3C7"  # Amarillo
            
            # ID Producto
            cell = self.ws.cell(row=row, column=1)
            cell.value = producto.get('idproducto', '')
            self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Nombre
            cell = self.ws.cell(row=row, column=2)
            cell.value = producto.get('nombre', '')
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            
            # Categoría
            cell = self.ws.cell(row=row, column=3)
            cell.value = producto.get('categoria', '')
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            
            # Stock
            cell = self.ws.cell(row=row, column=4)
            cell.value = stock
            color_texto = "DC2626" if stock == 0 else None
            self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_fondo, color_texto=color_texto)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Peso
            cell = self.ws.cell(row=row, column=5)
            cell.value = producto.get('peso', 0)
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Estante
            cell = self.ws.cell(row=row, column=6)
            cell.value = producto.get('id_estante', '')
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Estado
            cell = self.ws.cell(row=row, column=7)
            cell.value = estado
            self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_fondo)
            
            # Fecha Elaboración
            cell = self.ws.cell(row=row, column=8)
            fecha_elab = producto.get('fecha_elaboracion', '')
            if fecha_elab:
                cell.value = VencimientoHelper.formatear_fecha(fecha_elab)
            else:
                cell.value = "-"
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha Vencimiento
            cell = self.ws.cell(row=row, column=9)
            if fecha_venc:
                cell.value = VencimientoHelper.formatear_fecha(fecha_venc)
                # Color específico según estado de vencimiento
                if estado_vencimiento == 'vencido':
                    color_venc_fondo = "FCA5A5"  # Rojo
                    color_venc_texto = "991B1B"
                elif estado_vencimiento == 'vence_hoy':
                    color_venc_fondo = "FCA5A5"  # Rojo
                    color_venc_texto = "991B1B"
                elif estado_vencimiento == 'critico':
                    color_venc_fondo = "FDBA74"  # Naranja
                    color_venc_texto = "9A3412"
                elif estado_vencimiento == 'proximo':
                    color_venc_fondo = "FDE68A"  # Amarillo
                    color_venc_texto = "92400E"
                else:
                    color_venc_fondo = color_fondo
                    color_venc_texto = None
                self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_venc_fondo, color_texto=color_venc_texto)
            else:
                cell.value = "-"
                self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Días Restantes
            cell = self.ws.cell(row=row, column=10)
            if dias_restantes is not None:
                if dias_restantes < 0:
                    cell.value = f"Vencido ({abs(dias_restantes)}d)"
                    color_dias = "DC2626"
                elif dias_restantes == 0:
                    cell.value = "HOY"
                    color_dias = "DC2626"
                elif dias_restantes <= 7:
                    cell.value = f"{dias_restantes} días"
                    color_dias = "EA580C"
                elif dias_restantes <= 30:
                    cell.value = f"{dias_restantes} días"
                    color_dias = "D97706"
                else:
                    cell.value = f"{dias_restantes} días"
                    color_dias = "059669"
                self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_fondo, color_texto=color_dias)
            else:
                cell.value = "-"
                self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha Ingreso
            cell = self.ws.cell(row=row, column=11)
            fecha_valor = producto.get('fecha_ingreso', '')
            if isinstance(fecha_valor, str) and fecha_valor:
                try:
                    fecha_dt = datetime.fromisoformat(fecha_valor.replace('Z', '+00:00'))
                    cell.value = fecha_dt.strftime('%d/%m/%Y')
                except:
                    cell.value = fecha_valor
            else:
                cell.value = fecha_valor
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        # Resumen estadístico (con mejor diseño)
        row += 2
        self.ws.merge_cells(f'A{row}:B{row}')
        resumen_cell = self.ws[f'A{row}']
        resumen_cell.value = "📊 RESUMEN ESTADÍSTICO"
        resumen_cell.font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
        resumen_cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        resumen_cell.alignment = Alignment(horizontal='center', vertical='center')
        resumen_cell.border = self._crear_borde()
        self.ws.row_dimensions[row].height = 28
        
        row += 1
        total_productos = len(productos)
        sin_stock = sum(1 for p in productos if p.get('stock', 0) == 0)
        peso_total = sum(p.get('peso', 0) * p.get('stock', 0) for p in productos)
        
        # Estadísticas de vencimiento
        productos_vencidos = sum(1 for p in productos 
                                if p.get('fecha_vencimiento') and 
                                VencimientoHelper.calcular_dias_hasta_vencimiento(p['fecha_vencimiento']) is not None and
                                VencimientoHelper.calcular_dias_hasta_vencimiento(p['fecha_vencimiento']) < 0)
        
        productos_criticos = sum(1 for p in productos 
                                if p.get('fecha_vencimiento') and 
                                VencimientoHelper.calcular_dias_hasta_vencimiento(p['fecha_vencimiento']) is not None and
                                0 <= VencimientoHelper.calcular_dias_hasta_vencimiento(p['fecha_vencimiento']) <= 7)
        
        estadisticas = [
            ('📦 Total de productos', total_productos, None),
            ('❌ Productos sin stock', sin_stock, "DC2626" if sin_stock > 0 else None),
            ('⚠️ Productos vencidos', productos_vencidos, "DC2626" if productos_vencidos > 0 else None),
            ('🔔 Vencen en 7 días o menos', productos_criticos, "F59E0B" if productos_criticos > 0 else None),
            ('⚖️ Peso total inventario (kg)', f"{peso_total:,.2f}", "10B981"),
        ]
        
        for stat_label, stat_value, color_valor in estadisticas:
            cell_label = self.ws.cell(row=row, column=1)
            cell_label.value = stat_label
            cell_label.font = Font(name='Calibri', size=11, bold=True, color="1F2937")
            cell_label.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell_label.border = self._crear_borde(style='thin')
            cell_label.alignment = Alignment(horizontal='left', vertical='center')
            
            cell_value = self.ws.cell(row=row, column=2)
            cell_value.value = stat_value
            cell_value.font = Font(name='Calibri', size=11, bold=True, color=color_valor if color_valor else "1F2937")
            cell_value.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell_value.alignment = Alignment(horizontal='center', vertical='center')
            cell_value.border = self._crear_borde(style='thin')
            self.ws.row_dimensions[row].height = 22
            row += 1
        
        # Ajustar anchos
        self.ws.column_dimensions['A'].width = 10  # ID
        self.ws.column_dimensions['B'].width = 35  # Nombre
        self.ws.column_dimensions['C'].width = 20  # Categoría
        self.ws.column_dimensions['D'].width = 10  # Stock
        self.ws.column_dimensions['E'].width = 12  # Peso
        self.ws.column_dimensions['F'].width = 10  # Estante
        self.ws.column_dimensions['G'].width = 12  # Estado
        self.ws.column_dimensions['H'].width = 15  # Fecha Elaboración
        self.ws.column_dimensions['I'].width = 15  # Fecha Vencimiento
        self.ws.column_dimensions['J'].width = 18  # Días Restantes
        self.ws.column_dimensions['K'].width = 15  # Fecha Ingreso (antes era H)
        
        self._agregar_pie_pagina()
        
        # Guardar en memoria
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    def exportar_ventas(self, ventas, filtros=None):
        """Exporta ventas a Excel con formato profesional"""
        self.ws.title = "Ventas"
        
        # Encabezado
        row = self._agregar_encabezado_principal(
            "SISTEMA WEIGENCE - REPORTE DE VENTAS",
            len(ventas),
            filtros
        )
        
        # Encabezados de columnas
        headers = ['ID Venta', 'Fecha', 'Vendedor', 'RUT Vendedor', 'Total ($)', 'Productos', 'Unidades Totales']
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_num)
            cell.value = header
            self._aplicar_estilo_header(cell, self.COLOR_SUBTITLE)
        
        row += 1
        
        # Datos de ventas
        for idx, venta in enumerate(ventas):
            color_fondo = self.COLOR_ALTERNADO if idx % 2 == 0 else None
            
            # ID Venta
            cell = self.ws.cell(row=row, column=1)
            cell.value = venta.get('id_venta', '')
            self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha
            cell = self.ws.cell(row=row, column=2)
            fecha_valor = venta.get('fecha', '')
            if isinstance(fecha_valor, str) and fecha_valor:
                try:
                    fecha_dt = datetime.fromisoformat(fecha_valor.replace('Z', '+00:00'))
                    cell.value = fecha_dt.strftime('%d/%m/%Y %H:%M')
                except:
                    cell.value = fecha_valor
            else:
                cell.value = fecha_valor
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Vendedor
            cell = self.ws.cell(row=row, column=3)
            cell.value = venta.get('vendedor_nombre', '')
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            
            # RUT Vendedor
            cell = self.ws.cell(row=row, column=4)
            cell.value = venta.get('vendedor_rut', '')
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Total
            cell = self.ws.cell(row=row, column=5)
            total = venta.get('total', 0)
            cell.value = float(total)
            cell.number_format = '$#,##0.00'
            self._aplicar_estilo_celda(cell, bold=True, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Productos
            cell = self.ws.cell(row=row, column=6)
            cell.value = venta.get('total_productos', 0)
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Unidades
            cell = self.ws.cell(row=row, column=7)
            cell.value = venta.get('total_unidades', 0)
            self._aplicar_estilo_celda(cell, color_fondo=color_fondo)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        # Resumen estadístico (con mejor diseño)
        row += 2
        self.ws.merge_cells(f'A{row}:B{row}')
        resumen_cell = self.ws[f'A{row}']
        resumen_cell.value = "💰 RESUMEN ESTADÍSTICO"
        resumen_cell.font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
        resumen_cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        resumen_cell.alignment = Alignment(horizontal='center', vertical='center')
        resumen_cell.border = self._crear_borde()
        self.ws.row_dimensions[row].height = 28
        
        row += 1
        total_ventas = len(ventas)
        total_general = sum(float(v.get('total', 0)) for v in ventas)
        promedio_venta = total_general / total_ventas if total_ventas > 0 else 0
        total_productos = sum(v.get('total_productos', 0) for v in ventas)
        total_unidades = sum(v.get('total_unidades', 0) for v in ventas)
        
        estadisticas = [
            ('🛒 Total de ventas', total_ventas, None),
            ('💵 Total general', f"${total_general:,.2f}", "10B981"),
            ('📊 Promedio por venta', f"${promedio_venta:,.2f}", "3B82F6"),
            ('📦 Total productos vendidos', total_productos, None),
            ('🔢 Total unidades vendidas', total_unidades, "F59E0B"),
        ]
        
        for stat_label, stat_value, color_valor in estadisticas:
            cell_label = self.ws.cell(row=row, column=1)
            cell_label.value = stat_label
            cell_label.font = Font(name='Calibri', size=11, bold=True, color="1F2937")
            cell_label.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell_label.border = self._crear_borde(style='thin')
            cell_label.alignment = Alignment(horizontal='left', vertical='center')
            
            cell_value = self.ws.cell(row=row, column=2)
            cell_value.value = stat_value
            cell_value.font = Font(name='Calibri', size=11, bold=True, color=color_valor if color_valor else "1F2937")
            cell_value.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell_value.alignment = Alignment(horizontal='center', vertical='center')
            cell_value.border = self._crear_borde(style='thin')
            self.ws.row_dimensions[row].height = 22
            row += 1
        
        # Ajustar anchos
        self.ws.column_dimensions['A'].width = 12  # ID
        self.ws.column_dimensions['B'].width = 18  # Fecha
        self.ws.column_dimensions['C'].width = 25  # Vendedor
        self.ws.column_dimensions['D'].width = 15  # RUT
        self.ws.column_dimensions['E'].width = 15  # Total
        self.ws.column_dimensions['F'].width = 12  # Productos
        self.ws.column_dimensions['G'].width = 15  # Unidades
        
        self._agregar_pie_pagina()
        
        # Guardar en memoria
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
